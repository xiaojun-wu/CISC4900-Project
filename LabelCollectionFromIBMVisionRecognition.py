import openpyxl
import json
import os
from watson_developer_cloud import VisualRecognitionV3

"""
Initialize the API by API's key.
"""
visual_recognition = VisualRecognitionV3(
    '2018-03-19',
    iam_apikey='NUz9EQT9ZZDdhMUQLSx2m9awBSUEQ5_ztpabzf3xpquY')

"""
keywords is target label from the previous label recognition by myself, and it uses to filter the labels from API.
"""
keywords = ['animal','mammal','ice bear','bear','carnivore','dall sheep','wild sheep','ruminant','polar hare',
'hare','gnawing mammal','great white heron','heron','aquatic']

"""
dicts is used to store all labels from API, it's unfitlered labels.
description is string format of the filtered labels.
keys is string format of the unfiltered labels.
"""
dicts = {}
description = ''
keys = ''

"""
Create a workbook and read the excel file.
Get total rows from excel file.
"""
workbook = openpyxl.load_workbook('imageDescriptionByIBM.xlsx')
sheet = workbook.active

rows = sheet.max_row

"""
This function is to determine the label is valid or not.
"""
def isMatchKeyWords(key):
  return any(word == key.lower() for word in keywords)

"""
Read image name from excel file.
Concatenate the file path and file name.
Get response from Clarifai vision recognition API by file name and bucket name.
Save all the labels in dicts variable.
Filter the labels and save to description variable.
"""
for i in range(111,rows+1):
    filename = sheet.cell(row = i, column = 1).value
    # paht is absoluted path of folder that contains images.
    path = 'G:/CISC4900/ProjectImages/HITS/'
    fullpath = os.path.join(path,filename)
    fullpath = fullpath.replace('\\','')

    with open(fullpath, 'rb') as images_file:
        classes = visual_recognition.classify(
            images_file,
            threshold='0.0',
        	classifier_ids='default').get_result()
    classes = classes['images'][0]['classifiers'][0]['classes']
    for clas in classes:
        dicts[clas['class']]  = clas['score']
        print(clas['class'], clas['score'])

    for key in dicts:
        keys = keys + key + ','
        if(isMatchKeyWords(key)):
            description = description + key + ','

"""
Delete the last comma in description and keys.
If the description is empty, chose the higtest score label and fill in it.
"""
    description = description[:-1]
    keys = keys[:-1]

    if(description == ''):
        description = max(dicts,key = dicts.get)

"""
Save the description and keys in excel file.
They are filtered labels and unfiltered labels.
empty the variables for next loop.
"""
    sheet.cell(row = i, column = 2).value = description
    sheet.cell(row = i, column = 3).value = keys

    dicts = {}
    description = ''
    keys = ''

workbook.save("imageDescriptionByIBM.xlsx")
