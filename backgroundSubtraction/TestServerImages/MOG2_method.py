from __future__ import print_function
import cv2 as cv
import argparse
import openpyxl
import os

def main():
    #read image's name from excel file.
    workbook = openpyxl.load_workbook('fileNameInServerImages.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    backSub = cv.createBackgroundSubtractorMOG2()
    fileDir = os.path.dirname(os.path.abspath(__file__))
    parentDir = os.path.dirname(fileDir)

    for i in range(2,rows+1):
        filename = sheet.cell(row = i, column = 1).value
        path = os.path.join(parentDir,"serverImages/first1000In12-100RECNX/")
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','/')

        frame = cv.imread(fullpath)

        if frame is None:
            print("Can't find path.")
            break
        fgMask = backSub.apply(frame)
        res = cv.bitwise_and(frame,frame,mask = fgMask)
        filtered_image = filter_mask(res)
        # th_filtered_image = filter_mask_by_threshold(gray_res)
        
        cv.imshow('Frame', frame) #origin image
        cv.imshow('fgMask', fgMask) # foreground mask
        cv.imshow('FG Object',res) # color foreground
        cv.imshow('filtered_FG', filtered_image) # filtered color foreground

        #save the foreground into folder.

        path = os.path.join(fileDir,'mask/')
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','/')
        cv.imwrite(fullpath, fgMask)

        path = os.path.join(fileDir,'foregorund/')
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','/')
        cv.imwrite(fullpath, res)

        path = os.path.join(fileDir,'filtered_foreground/')
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','/')
        cv.imwrite(fullpath, filtered_image)
        
        keyboard = cv.waitKey(200)
        if keyboard == 'q' or keyboard == 27:
            break

def filter_mask(image):
    kernel = cv.getStructuringElement(cv.MORPH_CROSS,(2,2))

    closing = cv.morphologyEx(image, cv.MORPH_CLOSE, kernel)
    opening = cv.morphologyEx(closing, cv.MORPH_OPEN, kernel)

    return opening

def filter_mask_by_threshold(image):
    return cv.adaptiveThreshold(image,255,cv.ADAPTIVE_THRESH_GAUSSIAN_C,cv.THRESH_BINARY,11,2)

if __name__=='__main__':
    main()