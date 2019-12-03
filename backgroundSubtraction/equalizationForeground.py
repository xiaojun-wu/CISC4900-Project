from __future__ import print_function
import cv2 as cv
import argparse
import openpyxl
import os

def getFGfromEqualizedImage():
    workbook = openpyxl.load_workbook('filenames.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    backSub = cv.createBackgroundSubtractorMOG2()

    for i in range(2,rows+1):
        filename = sheet.cell(row = i, column = 1).value
        path = 'G:/CISC4900/backgroundSubtraction/equalization/'
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','')
        print(fullpath)

        frame = getFrame(fullpath)
        fgMask = backSub.apply(frame)
        res = cv.bitwise_and(frame,frame,mask = fgMask)
        filteredImage = filter_mask(res)
        cv.imshow('fg',filteredImage)
        keyboard = cv.waitKey(300)

        path = 'G:/CISC4900/backgroundSubtraction/equalization/filterImages/'
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','')
        cv.imwrite(fullpath, filteredImage)

def getFGfromClaheImage():
    workbook = openpyxl.load_workbook('filenames.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    backSub = cv.createBackgroundSubtractorMOG2()

    for i in range(2,rows+1):
        filename = sheet.cell(row = i, column = 1).value
        path = 'G:/CISC4900/backgroundSubtraction/CLAHE/'
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','')
        print(fullpath)

        frame = getFrame(fullpath)
        fgMask = backSub.apply(frame)
        res = cv.bitwise_and(frame,frame,mask = fgMask)
        filteredImage = filter_mask(res)
        cv.imshow('fg',filteredImage)
        keyboard = cv.waitKey(300)

        path = 'G:/CISC4900/backgroundSubtraction/CLAHE/foregrounds/'
        fullpath = os.path.join(path,filename)
        fullpath = fullpath.replace('\\','')
        cv.imwrite(fullpath, filteredImage)

def getFrame(path):
    frame = cv.imread(path,0)
    return frame

def filter_mask(image):
    kernel = cv.getStructuringElement(cv.MORPH_CROSS,(2,2))

    closing = cv.morphologyEx(image, cv.MORPH_CLOSE, kernel)
    opening = cv.morphologyEx(closing, cv.MORPH_OPEN, kernel)

    return opening

if __name__=='__main__':
    getFGfromEqualizedImage()
    getFGfromClaheImage()