import numpy as np
import cv2 as cv
from matplotlib import pyplot as plt
import openpyxl
import os

def getImagePath(imagName):
    path = 'G:/CISC4900/backgroundSubtraction/photos/'
    fullpath = os.path.join(path,imagName)
    fullpath = fullpath.replace('\\','')
    return fullpath

# This function is showing the histogram of each images in array.
def makeHistogram(nameArray):
    colors = ['b','g','r','c','m','y','k','w']

    for name,i in zip(nameArray, range(len(nameArray))):
        path = getImagePath(name)
        frame = cv.imread(path)
        hist = cv.calcHist([frame],[0],None,[256],[0,256])
        plt.plot(hist,color = colors[i])

    plt.show()

# equalized images, the input images should be grayscale
def equalizaiton():
    workbook = openpyxl.load_workbook('filenames.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    for i in range(2,rows+1):
        name = sheet.cell(row = i, column = 1).value
        path = getImagePath(name)
        frame = cv.imread(path,0)
        eFrame = cv.equalizeHist(frame)

        cv.imshow('equalization',eFrame)
        keyboard = cv.waitKey(300)

        path = 'G:/CISC4900/backgroundSubtraction/equalization/'
        fullpath = os.path.join(path,name)
        fullpath = fullpath.replace('\\','')
        cv.imwrite(fullpath, eFrame)

# CLAHE (Contrast Limited Adaptive Histogram Equalization)
def CLAHE():
    workbook = openpyxl.load_workbook('filenames.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    for i in range(2,rows+1):
        name = sheet.cell(row = i, column = 1).value
        path = getImagePath(name)
        frame = cv.imread(path,0)
        clahe = cv.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))    
        cl1 = clahe.apply(frame)

        cv.imshow('equalization',cl1)
        # keyboard = cv.waitKey(300)

        path = 'G:/CISC4900/backgroundSubtraction/CLAHE/'
        fullpath = os.path.join(path,name)
        fullpath = fullpath.replace('\\','')
        cv.imwrite(fullpath, cl1)

# equalized the color images.
def hisEqulColor(img):
    ycrcb = cv.cvtColor(img, cv.COLOR_BGR2YCR_CB)
    channels = cv.split(ycrcb)
    print (len(channels))
    cv.equalizeHist(channels[0], channels[0])
    cv.merge(channels, ycrcb)
    cv.cvtColor(ycrcb, cv.COLOR_YCR_CB2BGR, img)
    return img

if __name__=='__main__':
    names = ['RCNX4714.JPG','RCNX4715.JPG','RCNX4716.JPG','RCNX4717.JPG']
    #makeHistogram(names)
    equalizaiton()
    CLAHE()