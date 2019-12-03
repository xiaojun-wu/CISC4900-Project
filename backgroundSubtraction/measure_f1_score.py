import os
import openpyxl
import operator

global TP_dict
global FN_dict
global FP_dict
global TN_dict
global precision_dict
global recall_dict
global f1_score_dict

TP_dict = dict()
FN_dict = dict()
FP_dict = dict()
TN_dict = dict()
precision_dict = dict()
recall_dict = dict()
f1_score_dict = dict()

def imageGroundTruthArray():
    workbook = openpyxl.load_workbook('filtered_image_labels.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_true = []
    for i in range(2,rows+1):
        isCaribou = sheet.cell(row = i, column = 2).value
        isCaribou.upper()
        if (isCaribou == 'YES'):
            y_true.append(1)
        else:
            y_true.append(0)

    return y_true

def GoogleLabelIndicator(col):
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    firstLevelKeywords = ['deer','caribou','reindeer']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    workbook = openpyxl.load_workbook('filtered_image_labels.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_pred = []

    for i in range(2,rows+1):
        labelAnnotations = sheet.cell(row = i, column = col).value
        labels = list(labelAnnotations.split(","))

        value = 0
        for label in labels:
            if (any(word == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

def GoogleObjectDetector(col):
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    firstLevelKeywords = ['deer','caribou','reindeer']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    workbook = openpyxl.load_workbook('filtered_image_labels.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_pred = []

    for i in range(2,rows+1):
        objectAnnotations = sheet.cell(row = i, column = col).value
        labels = []
        if objectAnnotations is not None:
            objectLabels = list(objectAnnotations.split(","))
        else:
            objectLabels = []

        value = 0
        for label in objectLabels:
            if (any(word == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

def confusionMatrixAndDataDictGenerator(y_true,y_pred,name):
    TP=0
    FP=0
    FN=0
    TN=0
    for i in range(0,len(y_true)):
        if(y_true[i] == 1):
            TP += y_pred[i]
            FN += (1 - y_pred[i])
        else:
            FP += y_pred[i]
            TN += (1 - y_pred[i])

    print("TP is: ", TP),
    print("FN is: ", FN),
    print("FP is: ", FP),
    print("TN is: ", TN),

    precision = TP/(TP+FP)
    recall = TP/(TP+FN)

    print("precision is: ", precision),
    print("reall is: ", recall),

    f1_score = 2*((precision*recall)/(precision+recall))

    print("f1_score is: ", f1_score),
    print("\n")

    TP_dict[name] = TP
    FN_dict[name] = FN
    FP_dict[name] = FP
    TN_dict[name] = TN
    precision_dict[name] = precision
    recall_dict[name] = recall
    f1_score_dict[name] = f1_score

    return f1_score

y_true = imageGroundTruthArray()

y_pred = GoogleLabelIndicator(3)
print("Original images F1 measure result of Google cloudVision Label detect API:")
Google_F1_score_without_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithoutObj')

y_pred = GoogleObjectDetector(5)
print("Original images F1 measure result of Google cloudVision Object detection API:")
Google_F1_score_with_only_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithOnlyObj')

y_pred = GoogleLabelIndicator(4)
print("foreground images F1 measure result of Google cloudVision Label detect API:")
Google_F1_score_without_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithoutObj')

y_pred = GoogleObjectDetector(6)
print("foreground images F1 measure result of Google cloudVision Object detection API:")
Google_F1_score_with_only_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithOnlyObj')