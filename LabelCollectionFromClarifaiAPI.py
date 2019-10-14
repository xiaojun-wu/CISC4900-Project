from clarifai.rest import ClarifaiApp
import os
import openpyxl
"""
Initialize the API by API's key.
"""
app = ClarifaiApp(api_key='a84e8ec4a00942e79136f367442fe77e')

model = app.public_models.general_model

"""
keywords is target label from the previous label recognition by myself, and it uses to filter the labels from API.
"""
keywords = ['wildlife','animal','mammal','deer','caribou']

"""
dicts is used to store all labels from API, it's unfitlered labels.
description is string format of the filtered labels.
keys is string format of the unfiltered labels.
"""
dicts = {}
description = ''
keys = ''

"""
Create a workbook and read the excel file.
Get total rows from excel file.
"""
workbook = openpyxl.load_workbook('imageDescriptionByClarifai.xlsx')
sheet = workbook.active

rows = sheet.max_row

"""
This function is to determine the label is valid or not.
"""
def isMatchKeyWords(key):
  return any(word == key.lower() for word in keywords)

"""
Read image name from excel file.
Concatenate the file path and file name.
Get response from Clarifai vision recognition API by file name and bucket name.
Save all the labels in dicts variable.
Filter the labels and save to description variable.
"""
for i in range(105,rows+1):
    filename = sheet.cell(row = i, column = 1).value
    path = 'D:/CISC4900/ProjectImages/HITS/'
    fullpath = os.path.join(path,filename)
    fullpath = fullpath.replace('\\','')

    response = model.predict_by_filename(fullpath)

    concepts = response['outputs'][0]['data']['concepts']
    for concept in concepts:
        dicts[concept['name']]  = concept['value']

    for key in dicts:
        keys = keys + key + ','
        if(isMatchKeyWords(key)):
            description = description + key + ','

"""
Delete the last comma in description and keys.
If the description is empty, chose the higtest score label and fill in it.
"""
    description = description[:-1]
    keys = keys[:-1]

    if(description == ''):
        description = max(dicts,key = dicts.get)

"""
Save the description and keys in excel file.
They are filtered labels and unfiltered labels.
empty the variables for next loop.
"""
    sheet.cell(row = i, column = 2).value = description
    sheet.cell(row = i, column = 3).value = keys

    dicts = {}
    description = ''
    keys = ''
workbook.save("imageDescriptionByClarifai.xlsx")
