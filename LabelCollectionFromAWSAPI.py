import boto3
import os
import openpyxl
"""
keywords is target label from the previous label recognition by myself, and it uses to filter the labels from API.
"""
keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
'Arctic Fox','Mountain Goat','Bison']

"""
dicts is used to store all labels from API, it's unfitlered labels.
description is string format of the filtered labels.
keys is string format of the unfiltered labels.
"""
dicts = {}
description = ''
keys = ''

"""
Create a workbook and read the excel file.
Get total rows from excel file.
"""
workbook = openpyxl.load_workbook('imageDescriptionByAWS2.xlsx')
sheet = workbook.active

rows = sheet.max_row

"""
This function is to determine the label is valid or not.
"""
def isMatchKeyWords(key):
  return any(word == key.lower() for word in keywords)

"""
bucket name from my AWS S3 storage.
"""
bucket='cisc4900project'

client=boto3.client('rekognition','us-east-2')

"""
Read image name from excel file.
Get response from AWS vision recognition API by file name and bucket name.
Save all the labels in dicts variable.
Filter the labels and save to description variable.
"""
for i in range(106,rows+1):
    filename = sheet.cell(row = i, column = 1).value

    response = client.detect_labels(Image={'S3Object':{'Bucket':bucket,'Name':filename}},MaxLabels=20,MinConfidence=0)

    labels = response['Labels']
    for label in labels:
        dicts[label['Name']]  = label['Confidence']

    for key in dicts:
        keys = keys + key + ','
        if(isMatchKeyWords(key)):
            description = description + key + ','
"""
Delete the last comma in description and keys.
If the description is empty, chose the higtest score label and fill in it.
"""
    description = description[:-1]
    keys = keys[:-1]

    if(description == ''):
        description = max(dicts,key = dicts.get)
"""
Save the description and keys in excel file.
They are filtered labels and unfiltered labels.
empty the variables for next loop.
"""
    sheet.cell(row = i, column = 2).value = description
    sheet.cell(row = i, column = 3).value = keys

    dicts = {}
    description = ''
    keys = ''

workbook.save("imageDescriptionByAWS2.xlsx")
