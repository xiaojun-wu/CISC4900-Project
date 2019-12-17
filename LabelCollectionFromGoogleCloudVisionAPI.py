import io
import os
import openpyxl

# Imports the Google Cloud client library
from google.cloud import vision
from google.cloud.vision import types
"""
Provide the credential
"""
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "G:/CISC4900/CISC 4900 Project-ce4a6163112b.json"

"""
keywords is target label from the previous label recognition by myself, and it uses to filter the labels from API.
"""
keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
'Arctic Fox','Mountain Goat','Bison']

"""
dicts is used to store all labels from detect_label API, it's unfitlered labels.
objectDetectedDict is used to store all labels from detec_object API, it's unfitlered labels.
description is string format of the filtered labels from detect_label.
objects is string fromat varibale of the filtered lables from detect_object.
keys is string format of the unfiltered labels.
"""
dicts = {}
objectDetectedDict = {}
description = ''
objects = ''
keys = ''

"""
Create a workbook and read the excel file.
Get total rows from excel file.
"""
workbook = openpyxl.load_workbook('imageDescriptionByCloudVisionAPIs.xlsx')
sheet = workbook.active

rows = sheet.max_row

"""
This function is to determine the label is valid or not.
"""
def isMatchKeyWords(key):
  return any(word == key.lower() for word in keywords)

# Instantiates a client
client = vision.ImageAnnotatorClient()

"""
Read image name from excel file.
Concatenate the file path and file name.
Get response from label_detection API and object_dectection API by file name and bucket name.
Save label-detection results in dicts variable.
Filter the labels and save to description variable.
Save object_detection results in objectDetectedDict variable.
Fileter the labels and save to objects variable.
"""
for i in range(151,rows+1):
    filename = sheet.cell(row = i, column = 1).value
    # path is the cloud storage link that contains images.
    path = 'gs://cisc4900project/HITS/'
    fullpath = os.path.join(path,filename)
    fullpath = fullpath.replace('\\','')

    response = client.annotate_image({
       'image': {'source': {'image_uri': fullpath}},
       'features': [{'type': vision.enums.Feature.Type.LABEL_DETECTION,'max_results': 20},
                    {'type': vision.enums.Feature.Type.OBJECT_LOCALIZATION}],
    })

    labels = response.label_annotations
    for label in labels:
        dicts[label.description]  = label.score

    for key in dicts:
        keys = keys + key + ','
        if(isMatchKeyWords(key)):
            description = description + key + ','

    labels = response.localized_object_annotations
    for label in labels:
        objectDetectedDict[label.name] = label.score

    for key in objectDetectedDict:
        keys = keys + key + ','
        objects = objects + key + ','
"""
Delete the last comma in description, keys and objects.
If the description is empty, chose the higtest score label and fill in it.
"""
    description = description[:-1]
    keys = keys[:-1]
    objects = objects[:-1]

    if(description == ''):
        description = max(dicts,key = dicts.get)
"""
Save the description,keys and objects in excel file.
They are filtered labels and unfiltered labels.
empty the variables for next loop.
"""
    sheet.cell(row = i, column = 2).value = description
    sheet.cell(row = i, column = 3).value = objects
    sheet.cell(row = i, column = 4).value = keys

    dicts = {}
    objectDetectedDict = {}
    description = ''
    keys = ''
    objects = ''


workbook.save("imageDescriptionByCloudVisionAPIs.xlsx")
