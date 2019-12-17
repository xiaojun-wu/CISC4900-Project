import os
import PIL.Image
import openpyxl

'''
This program is going to create a spreadsheet that contains all image's filename.
First it create a spreadsheet and name it 'sample_book'.
Second set first cell is 'filename'
Then read image's name in certain folder, and save them in spreadsheet.
The image's name are below the first cell and one by one.
So the first column in spreadsheet are file names.
'''
wb = openpyxl.Workbook()
wb.save(filename = 'sample_book.xlsx')
sheet = wb.active

sheet.cell(row=1,column=1).value = "fileName"
row = 2

with os.scandir('ProjectImages/HITS/') as entries:
    for entry in entries:
        print(entry.name)
        sheet.cell(row = row,column=1).value = entry.name
        row = row+1

wb.save('sample_book.xlsx')

