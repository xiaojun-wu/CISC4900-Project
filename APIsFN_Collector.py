import os
import openpyxl

'''
This prograom is the filter program that select all FN images that they has caribou in it but API said no.
If the images was detect FN, mark the FN in spreadsheet and save the date.
Those date for analyze the reason of why API detect those image as FN.
'''

global workbook
global sheet
global rows

workbook = openpyxl.load_workbook('pattern/CollectionOfAPIsFN.xlsx')
sheet = workbook.active
rows = sheet.max_row

# Get the ground Truth array of each labels.
def imageGroundTruthArray():
    y_true = []
    for i in range(2,rows+1):
        isCaribou = sheet.cell(row = i, column = 2).value
        isCaribou.upper()
        if (isCaribou == 'YES'):
            y_true.append(1)
        else:
            y_true.append(0)

    return y_true

# Get Google label detection API's result.
# For each image, 0 means API doesn't detect any caribou, more that zero means detect caribous
# The score depend on how exact label descript the caribous.
def GoogleLabelFN():
    # labels considered API detect caribou in image.
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    # Most exact labels
    firstLevelKeywords = ['deer','caribou','reindeer']
    # Second exact labels
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    # third exact labels
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    y_pred = []

    # Create sheet
    sheet = workbook["GoogleLabel"]
    rows = sheet.max_row

    # Check if API detect caribou in image or not, and assign score on it.
    for i in range(2,rows+1):
        labelAnnotations = sheet.cell(row = i, column = 3).value
        labels = list(labelAnnotations.split(","))

        value = 0
        for label in labels:
            if (any(word == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    FailNegativeArray(sheet,y_pred)

# Get Google object detection API's result.
# For each image, 0 means API doesn't detect any caribou, more that zero means detect caribous
# The score depend on how exact label descript the caribous.
def GoogleOjectFN():
    # labels considered API detect caribou in image.
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    # Most exact labels
    firstLevelKeywords = ['deer','caribou','reindeer']
    # Second exact labels
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    # third exact labels
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    # Create sheet
    sheet = workbook["GoogleObject"]
    rows = sheet.max_row

    y_pred = []

    # Check if API detect caribou in image or not, and assign score on it.
    for i in range(2,rows+1):
        objectAnnotations = sheet.cell(row = i, column = 3).value
        labels = []
        if objectAnnotations is not None:
            objectLabels = list(objectAnnotations.split(","))
        else:
            objectLabels = []

        value = 0
        for label in objectLabels:
            if (any(word.lower() == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word.lower() == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word.lower() == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)

    FailNegativeArray(sheet,y_pred)

# Get Clarifai label detection API's result.
# For each image, 0 means API doesn't detect any caribou, more that zero means detect caribous
# The score depend on how exact label descript the caribous.
def ClarifaiLabelsFN():
    # labels considered API detect caribou in image.
    keywords = ['wildlife','animal','mammal','deer','caribou']
    # Most exact labels
    firstLevelKeywords = ['deer','caribou']
    secondLevelKeywords = []
    # third exact labels
    thirdLevelKeywords = ['wildlife','animal','mammal']

    # Create sheet.
    sheet = workbook["Clarifai"]
    rows = sheet.max_row

    y_pred = generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords)
    FailNegativeArray(sheet,y_pred)

# Get IBM label detection API's result.
# For each image, 0 means API doesn't detect any caribou, more that zero means detect caribous
# The score depend on how exact label descript the caribous.
def IBMLabelFN():
    keywords = ['animal','mammal','ice bear','bear','carnivore','dall sheep','wild sheep','ruminant','polar hare',
    'hare','gnawing mammal','great white heron','heron','aquatic']
    firstLevelKeywords = ['caribou','reindeer','deer']
    secondLevelKeywords = ['dall sheep', 'wild sheep']
    thirdLevelKeywords = ['animal','mammal','ice bear','bear','carnivore','ruminant','polar hare','hare','gnawing mammal',]

    sheet = workbook["IBM"]
    rows = sheet.max_row

    y_pred = generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords)
    FailNegativeArray(sheet,y_pred)

# Get AWS label detection API's result.
# For each image, 0 means API doesn't detect any caribou, more that zero means detect caribous
# The score depend on how exact label descript the caribous.
def AWSLabelFN():
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison','Buffalo']
    firstLevelKeywords = ['deer','caribou']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine',
    'Arctic Fox','Bison','Buffalo']

    sheet = workbook["AWS"]
    rows = sheet.max_row

    y_pred = generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords)
    FailNegativeArray(sheet,y_pred)

# Check if API detect caribou in image or not, and assign score on it, return array of prediction.
def generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords):
    y_pred = []

    for i in range(2,rows+1):
        inputs = sheet.cell(row = i, column = 3).value
        labels = list(inputs.split(","))

        value = 0
        for label in labels:
            if (any(word.lower() == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word.lower() == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word.lower() == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

# Compare Ground Truth and API's prediction, If labels if FN(ground Truth is yes, but API said no.)
# Mark FN.
def FailNegativeArray(sheet,y_pred):
    y_true = imageGroundTruthArray()
    count = 0

    for i in range(2,rows+1):
        if y_true[i-2] == 1 and y_pred[i-2] != 1 :
            count = count + 1
            sheet.cell(row = i, column = 4).value = "FN"
            sheet.cell(row = i, column = 5).value = y_pred[i-2]
        sheet.cell(row = 1, column = 8).value = "NumberOfFN"
        sheet.cell(row = 1, column = 9).value = count

    workbook.save("pattern/CollectionOfAPIsFN.xlsx")

GoogleLabelFN()
GoogleOjectFN()
ClarifaiLabelsFN()
IBMLabelFN()
AWSLabelFN()
