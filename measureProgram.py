import os
import openpyxl
import operator

"""
This program is tend to calculate the F1-score of 4 vision recognition APIs performence.
In the previous steps, I already collect the return labels from these APIs by dectecing same bunch of iamges.
The labels are saving in 5 excel files include the ground truth image information.
This program is going to read the excel file and extract the labels from it.
Analyze the labels, if there is a valid labels appear, give some credit in it.

I divide the labels into 4 parts:
First part is the labels descript the caribou perfectly, like "caribou", "reindeer", and "derr".
Second part is the labels is not exact descript the caribou but they are looks like caribou, like "sheep", "goat".
Third part is the labels are not specific like "animal", "mammal".
Fourth part is invalid labels, they are have no credit.

The program read the iamge info and API's info, generate the indicator array.
Compare two arrays and make a confusion matrix.
Than use the confusion matrix to calculate the F1-score for each API.

Since the Google vision recognition provide two useful APIs, the label detector and object detector.
I write three functions, first just count the label detector's return labels,
second one only count the object detector's return labels,
third one count both detector's result.

I also save the TN, FP, FN, TN, precision, recall, and f1-score in a dictionary and sort them with descending order.
It can see the different of these API's performence in each field.
"""

global TP_dict
global FN_dict
global FP_dict
global TN_dict
global precision_dict
global recall_dict
global f1_score_dict

TP_dict = dict()
FN_dict = dict()
FP_dict = dict()
TN_dict = dict()
precision_dict = dict()
recall_dict = dict()
f1_score_dict = dict()


def imageGroundTruthArray():
    workbook = openpyxl.load_workbook('image-descriptionByMyself.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_true = []
    for i in range(2,rows+1):
        isCaribou = sheet.cell(row = i, column = 2).value
        isCaribou.upper()
        if (isCaribou == 'YES'):
            y_true.append(1)
        else:
            y_true.append(0)

    return y_true

def IBMLabelIndicatorArray():
    keywords = ['animal','mammal','ice bear','bear','carnivore','dall sheep','wild sheep','ruminant','polar hare',
    'hare','gnawing mammal','great white heron','heron','aquatic']
    firstLevelKeywords = ['caribou','reindeer','deer']
    secondLevelKeywords = ['dall sheep', 'wild sheep']
    thirdLevelKeywords = ['animal','mammal','ice bear','bear','carnivore','ruminant','polar hare','hare','gnawing mammal',]

    workbook = openpyxl.load_workbook('imageDescriptionByIBM.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    return generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords)

def GoogleLabelIndicatorArrayNoObjectAnnoation():
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    firstLevelKeywords = ['deer','caribou','reindeer']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    workbook = openpyxl.load_workbook('imageDescriptionByCloudVisionAPIs.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_pred = []

    for i in range(3,rows+1):
        labelAnnotations = sheet.cell(row = i, column = 2).value
        labels = list(labelAnnotations.split(","))

        value = 0
        for label in labels:
            if (any(word == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

def GoogleLabelIndicatorArrayWithOnlyObjectDetector():
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    firstLevelKeywords = ['deer','caribou','reindeer']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    workbook = openpyxl.load_workbook('imageDescriptionByCloudVisionAPIs.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_pred = []

    for i in range(3,rows+1):
        objectAnnotations = sheet.cell(row = i, column = 3).value
        labels = []
        if objectAnnotations is not None:
            objectLabels = list(objectAnnotations.split(","))
        else:
            objectLabels = []

        value = 0
        for label in objectLabels:
            if (any(word == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

def GoogleLabelIndicatorArrayWithBothDetectors():
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison']
    firstLevelKeywords = ['deer','caribou','reindeer']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine','Arctic Fox','Bison']

    workbook = openpyxl.load_workbook('imageDescriptionByCloudVisionAPIs.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    y_pred = []

    for i in range(3,rows+1):
        labelAnnotations = sheet.cell(row = i, column = 2).value
        objectAnnotations = sheet.cell(row = i, column = 3).value
        labels = list(labelAnnotations.split(","))
        if objectAnnotations is not None:
            objectLabels = list(objectAnnotations.split(","))
        else:
            objectLabels = []
        labels = labels + objectLabels

        value = 0
        for label in labels:
            if (any(word == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

def ClarifaiLabelIndicatorArray():
    keywords = ['wildlife','animal','mammal','deer','caribou']
    firstLevelKeywords = ['deer','caribou']
    secondLevelKeywords = []
    thirdLevelKeywords = ['wildlife','animal','mammal']

    workbook = openpyxl.load_workbook('imageDescriptionByClarifai.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    return generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords)

def AWSLabelIndicatorArray():
    keywords = ['wildlife','animal','mammal','deer','caribou','bear','polar bear','antelope','horse','goat','sheep','Canine',
    'Arctic Fox','Mountain Goat','Bison','Buffalo']
    firstLevelKeywords = ['deer','caribou']
    secondLevelKeywords = ['antelope','horse','goat','sheep','Mountain Goat']
    thirdLevelKeywords = ['wildlife','animal','mammal','bear','polar bear','Canine',
    'Arctic Fox','Bison','Buffalo']

    workbook = openpyxl.load_workbook('imageDescriptionByAWSAPI.xlsx')
    sheet = workbook.active
    rows = sheet.max_row

    return generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords)

def generatePredictArray(sheet,rows,firstLevelKeywords,secondLevelKeywords,thirdLevelKeywords):
    y_pred = []

    for i in range(2,rows+1):
        inputs = sheet.cell(row = i, column = 2).value
        labels = list(inputs.split(","))

        value = 0
        for label in labels:
            if (any(word.lower() == label.lower() for word in firstLevelKeywords)):
                if value < 1:
                    value = 1
            elif (any(word.lower() == label.lower() for word in secondLevelKeywords)):
                if value < 0.8:
                    value = 0.8
            elif (any(word.lower() == label.lower() for word in thirdLevelKeywords)):
                if value < 0.4:
                    value = 0.4
        y_pred.append(value)
    return y_pred

def confusionMatrix(y_true,y_pred):
    TP=0
    FP=0
    FN=0
    TN=0
    for i in range(0,len(y_true)):
        if(y_true[i] == 1):
            TP += y_pred[i]
            FN += (1 - y_pred[i])
        else:
            FP += y_pred[i]
            TN += (1 - y_pred[i])

    print("TP is: ", TP),
    print("FN is: ", FN),
    print("FP is: ", FP),
    print("TN is: ", TN),

    precision = TP/(TP+FP)
    recall = TP/(TP+FN)

    print("precision is: ", precision),
    print("reall is: ", recall),

    f1_score = 2*((precision*recall)/(precision+recall))

    print("f1_score is: ", f1_score),
    print("\n")

    return f1_score

def confusionMatrixAndDataDictGenerator(y_true,y_pred,name):
    TP=0
    FP=0
    FN=0
    TN=0
    for i in range(0,len(y_true)):
        if(y_true[i] == 1):
            TP += y_pred[i]
            FN += (1 - y_pred[i])
        else:
            FP += y_pred[i]
            TN += (1 - y_pred[i])

    print("TP is: ", TP),
    print("FN is: ", FN),
    print("FP is: ", FP),
    print("TN is: ", TN),

    precision = TP/(TP+FP)
    recall = TP/(TP+FN)

    print("precision is: ", precision),
    print("reall is: ", recall),

    f1_score = 2*((precision*recall)/(precision+recall))

    print("f1_score is: ", f1_score),
    print("\n")

    TP_dict[name] = TP
    FN_dict[name] = FN
    FP_dict[name] = FP
    TN_dict[name] = TN
    precision_dict[name] = precision
    recall_dict[name] = recall
    f1_score_dict[name] = f1_score

    return f1_score

y_true = imageGroundTruthArray()
print("")

y_pred = IBMLabelIndicatorArray()
print("F1 measure result of IBM vision API:")
IBM_F1_score = confusionMatrixAndDataDictGenerator(y_true,y_pred,'IBM')

y_pred = ClarifaiLabelIndicatorArray()
print("F1 measure result of Clarifai vision API:")
Clarifai_F1_score = confusionMatrixAndDataDictGenerator(y_true,y_pred,'Clarifai')

y_pred = AWSLabelIndicatorArray()
print("F1 measure result of AWS vision API:")
AWS_F1_score = confusionMatrixAndDataDictGenerator(y_true,y_pred,'AWS')

y_pred = GoogleLabelIndicatorArrayNoObjectAnnoation()
print("F1 measure result of Google cloudVision API without object detection:")
Google_F1_score_without_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithoutObj')

y_pred = GoogleLabelIndicatorArrayWithOnlyObjectDetector()
print("F1 measure result of Google cloudVision API with only object detection:")
Google_F1_score_with_only_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithOnlyObj')

y_pred = GoogleLabelIndicatorArrayWithBothDetectors()
print("F1 measure result of Google cloudVision API with object detection:")
Google_F1_score_with_objectDetector = confusionMatrixAndDataDictGenerator(y_true,y_pred,'GoolgeWithBothDetector')

sorted_TP_dict = sorted(TP_dict.items(), key = operator.itemgetter(1), reverse = True)
sorted_FN_dict = sorted(FN_dict.items(), key = operator.itemgetter(1), reverse = True)
sorted_FP_dict = sorted(FP_dict.items(), key = operator.itemgetter(1), reverse = True)
sorted_TN_dict = sorted(TN_dict.items(), key = operator.itemgetter(1), reverse = True)
sorted_precision_dict = sorted(precision_dict.items(), key = operator.itemgetter(1), reverse = True)
sorted_recall_dict = sorted(recall_dict.items(), key = operator.itemgetter(1), reverse = True)
sorted_f1_score_dict = sorted(f1_score_dict.items(), key = operator.itemgetter(1), reverse = True)

print("After sorted, the descending order of TN, FN, FP, TN, precision, recall, and f1-score are:")
print("")
print("The TP result of APIs are:")
print(sorted_TP_dict,"\n")
print("The FN result of APIs are:")
print(sorted_FN_dict,"\n")
print("The FP result of APIs are:")
print(sorted_FP_dict,"\n")
print("The TN result of APIs are:")
print(sorted_TN_dict,"\n")
print("The precision result of APIs are:")
print(sorted_precision_dict,"\n")
print("The recall result of APIs are:")
print(sorted_recall_dict,"\n")
print("The f1-score result of APIs are:")
print(sorted_f1_score_dict,"\n")
